import urllib.request
import json
import pprint
import time
import math
import openpyxl
from openpyxl import Workbook
import datetime

#https://stackoverflow.com/questions/29824111/get-a-users-steam-inventory
def getInventory(steamid):
    data = urllib.request.urlopen('http://steamcommunity.com/inventory/'+steamid+'/753/6')
    json_data = json.loads(data.read())
    descriptions = json_data['descriptions']
    assets = json_data['assets']
    inv_count = json_data['total_inventory_count']
    print(inv_count)
    time = math.floor(float(inv_count)/25)*4 #25 items every 4 minutes
    print('this will take at least ' + str(time) + ' minutes')
        #print(item['market_hash_name'])
    writeToExcel(assets, descriptions)
    print ('Done!')
    #return

def writeToExcel(assets, descriptions):
    wb = Workbook() 
    ws = wb.active

    count = 0 #number of successful requests, should always reach 25
    r = 1
    col_name = 1
    col_price = 3
    col_game = 2
    ws.cell(r, col_game, 'Game')
    ws.cell(r, col_name, 'Name')
    ws.cell(r, col_name, 'Selling Price')
    r = r+1
    for item in descriptions: #item is a dict of one description
        tags = item['tags']
        if(len(tags) == 4):
            is_card = False
            game_tag = tags[1]
            class_tag = tags[3]
            if class_tag['localized_tag_name'] == 'Trading Card':
                is_card = True
            if is_card:
                game_name = game_tag['localized_tag_name']
                ws.cell(r, col_game, game_name)
                ws.cell(r, col_name, item['name'])
                
                price = getPrice(item['market_hash_name'], game_name, count)
                if(price == 0):
                    time.sleep(4*60)
                    while(price == 0):
                        price = getPrice(item['name'], game_name, count)
                    count = 0 
                count = count+1
                ws.cell(r, col_price, price)
                r = r+1


    wb.save('file.xlsx')

def getPrice(card, game, count):
    url = 'https://steamcommunity.com/market/search/render/?query='
    url = url+card+" "+game
    url = url.replace(" ", "+")
    print(url)
    t = datetime.datetime.now()
    passed = False
    try:
        data = urllib.request.urlopen(url)
        print("success")
        #print(data.getheaders())
        #k = data.headers.keys()
        json_data = json.loads(data.read())
        print(t)
        return parsePrice("hello")
    except Exception as e: #I need to make this for just HTTP errors
        print("requests exceeded: " + str(count))
        print(t)
        passed = True
        #time.sleep(1)
        return 0

def parsePrice(html):

    return 1
        
def start():
    try:
        #inp = input('please input steamID64')
        getInventory('76561198089894938')
    except:
        print('error')

start()#use __main__
