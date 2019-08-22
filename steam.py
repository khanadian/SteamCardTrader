import urllib.request
import json
import pprint
import time
import math
import openpyxl
from openpyxl import Workbook
import datetime
import queue

#this method is courtesy of https://stackoverflow.com/questions/29824111/get-a-users-steam-inventory
def getInventory(steamid):
    data = urllib.request.urlopen('http://steamcommunity.com/inventory/'+steamid+'/753/6')
    json_data = json.loads(data.read())
    descriptions = json_data['descriptions']
    assets = json_data['assets']
    writeToExcel(assets, descriptions)
    print ("Done!")
    #return


def writeToExcel(assets, descriptions):
    wb = Workbook() 
    ws = wb.active

    count = 0 #number of successful requests
    r = 1
    ws.cell(r, 2, 'Game')
    ws.cell(r, 1, 'Name')
    ws.cell(r, 3, 'Selling Price')
    r = r+1

    #queue must store item, game name, row number
    q = queue.Queue(0)
    q2 = queue.Queue(0)
    
    for item in descriptions: #item is a dict of one description
        tags = item['tags']
        if(len(tags) == 4):
            game_tag = tags[1]
            class_tag = tags[3]
            if class_tag['localized_tag_name'] == 'Trading Card':
                game_name = game_tag['localized_tag_name']
                count = checkPrice(ws, wb, item, game_name, count, q, r)
                r = r+1

    #go through queue of unsuccessful calls
    tup = q.get()
    curr_q = q
    curr_op = q2 #empty queues back and forth
                    #until list does not reduce or is empty
    l = 0 #length of the queue
    stalling = False
    attempts_allowed = 3 #number of time's we'll allow stalling
    attempts = 0
    while(not stalling):
        count = checkPrice(ws, wb, tup[0], tup[1], count, curr_op, tup[2])
        print("3")
        if(curr_q.empty()):
            print("4")
            tup = curr_op.get()
            curr_op, curr_q = curr_q, curr_op
            if (l == curr_q.qsize()):
                time.sleep(60)
                attempts = attempts + 1
                if (attempts > attempts_allowed):
                    stalling = True
                print("5")
            else:
                print("6")
                l = curr_q.qsize()
            print('length of queue = ' + str(l))
            print("7")
        else:
            tup = curr_q.get()
    
    print("8")

    #wb.save('file.xlsx')

def checkPrice(ws, wb, item, game_name, count, q, r):
    price = getPrice(item['market_hash_name'], game_name, count)
    print("2")
    if(price == 0):
        to_queue = (item, game_name, r)
        q.put(to_queue)
        count = 0
        print('added to queue')
    else:
        ws.cell(r, 2, game_name)
        ws.cell(r, 1, item['name'])
        ws.cell(r, 3, price)
        count = count+1
        wb.save('file.xlsx')

    return count

def getPrice(card, game, count):
    url = 'https://steamcommunity.com/market/priceoverview/?currency=20&appid=753&market_hash_name='
    url = url+card
    print(url)
    t = datetime.datetime.now()
    time.sleep(1)
    try:
        data = urllib.request.urlopen(url)
        print("success")
        #print(data.getheaders())
        #k = data.headers.keys()
        print(t)
        print("1")
        return parsePrice(data)
    except Exception as e: #I need to make this for just HTTP errors
        print("requests exceeded: " + str(count))
        print(t)
        #time.sleep(1)
        return 0

def parsePrice(data):
    json_data = json.loads(data.read())
    price = json_data['lowest_price']
    print(price)
    return price
        
def start():
    #inp = input('please input steamID64')
    getInventory('76561198089894938')

start()#use __main__
